import fitz  # PyMuPDF
import pandas as pd
import pytesseract
from pytesseract import Output
from PIL import Image
from io import BytesIO
import numpy as np
import os
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import warnings
from collections import defaultdict
import re
from pathlib import Path
import time
from openpyxl.styles import PatternFill

warnings.filterwarnings(
    "ignore",
    message=".*Unknown extension is not supported and will be removed.*",
    category=UserWarning,
)
pytesseract.pytesseract.tesseract_cmd = r"C:\Users\IlaBarshilia\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"

# ============================
# HIGHLIGHT CONFIG
# ============================

CHANGE_HIGHLIGHT_FILL = PatternFill(
    start_color="FFF2CC",  # light yellow (Excel-safe)
    end_color="FFF2CC",
    fill_type="solid"
)


all_input_filepath = input('Enter the path where all of your monthly input sources to update claims log are kept: ')
# --- Claims workbook (target) ---
Claims_filepath = input('Enter the path where Claims Log excel file is kept: ')
Claims_filename = input('Enter the name of Claims Log excel file without .xlsx: ') + ".xlsx"

# ============================
# TEXT-BASED PDF EXTRACTION
# ============================

def _clean_span_text(s: str) -> str:
    return (s or "").strip()

def extract_structured_text_textpdf(pdf_path: str) -> pd.DataFrame:
    """
    Extract text using the native PDF text layer.
    ALWAYS returns a DataFrame with a stable schema:
    ['page', 'text', 'x0', 'y0', 'x1', 'y1']
    """
    doc = fitz.open(pdf_path)
    rows = []

    for page_num, page in enumerate(doc, start=1):
        blocks = page.get_text("dict").get("blocks", [])
        for b in blocks:
            if "lines" not in b:
                continue
            for line in b["lines"]:
                for span in line["spans"]:
                    text = _clean_span_text(span.get("text", ""))
                    if not text:
                        continue
                    x0, y0, x1, y1 = map(round, span["bbox"])
                    rows.append({
                        "page": page_num,
                        "text": text,
                        "x0": x0, "y0": y0, "x1": x1, "y1": y1,
                    })

    return pd.DataFrame(rows, columns=["page", "text", "x0", "y0", "x1", "y1"])

def text_extraction_failed(df: pd.DataFrame) -> bool:
    """
    Safely determines whether a PDF has no usable text layer.
    """
    if df is None or df.empty:
        return True
    if not {"page", "text"}.issubset(df.columns):
        return True
    non_blank = df["text"].astype(str).str.strip().replace("", pd.NA).dropna()
    return len(non_blank) == 0

def group_by_dynamic_columns_textpdf(
    df: pd.DataFrame, y_tolerance: int = 2, x_bin_width: int = 20
) -> pd.DataFrame:
    """
    Column-aware grouping for TEXT PDFs.

    - Groups items into rows based on similar y0 (within y_tolerance).
    - Splits into columns using x0 binning (x_bin_width).
    """
    if df is None or df.empty:
        return pd.DataFrame()
    structured_rows = []

    for page in df["page"].unique():
        page_df = df[df["page"] == page]
        rows = []

        # 1) group spans into "lines" by y0 tolerance
        for _, r in page_df.iterrows():
            for row in rows:
                if abs(row["y0"] - r["y0"]) <= y_tolerance:
                    row["cells"].append(r)
                    break
            else:
                rows.append({"y0": r["y0"], "cells": [r]})

        # 2) within each line, bin by x0 into columns
        for row in rows:
            bins = defaultdict(list)
            for cell in row["cells"]:
                bin_idx = int(cell["x0"] // x_bin_width)
                bins[bin_idx].append(cell["text"])

            structured_rows.append([" ".join(bins[i]) for i in sorted(bins)])

    return pd.DataFrame(structured_rows)

# ============================
# ENTRY POINT (TEXT ONLY)
# ============================
def extract_pdf_text_only(pdf_path: str) -> pd.DataFrame:
    """
    Text-only pipeline:
      - extract native text layer
      - group into rows/columns
    """
    df_text = extract_structured_text_textpdf(pdf_path)
    if text_extraction_failed(df_text):
        # Text component should not OCR; caller decides what to do next
        return pd.DataFrame()
    return group_by_dynamic_columns_textpdf(df_text)

if __name__ == "__main__":
    united_path = next((p for p in Path(all_input_filepath).glob("*.pdf") if "lossrunsreport" in p.name.lower()), None)
    if united_path is None:
        print(f"No PDF containing 'LossRunsReport' found in: {all_input_filepath}")
        time.sleep(15)
        raise FileNotFoundError(f"No PDF containing 'LossRunsReport' found in: {all_input_filepath}")
    united_filename = united_path.name
    print("Found Loss Runs Report PDF: ", united_filename)
    united_pdf_path = os.path.join(all_input_filepath, united_filename)
    United_ins = extract_pdf_text_only(united_pdf_path)

def _is_str(x):
    return isinstance(x, str)

def _norm(s: str) -> str:
    return (s or "").strip().lower()

def _parse_money(x):
    if x is None:
        return None

    # numeric types
    if isinstance(x, (int, float, np.number)):
        if pd.isna(x):
            return None
        return float(x)

    if isinstance(x, str):
        s = x.strip()
        if s == "" or s.lower() == "none":
            return None

        # detect parentheses negative: (123.45) => -123.45
        neg = False
        if s.startswith("(") and s.endswith(")"):
            neg = True
            s = s[1:-1].strip()

        # remove common formatting
        s = s.replace("$", "").replace(",", "").strip()
        if s in ("", "-", "--"):
            return None

        try:
            val = float(s)
            return -val if neg else val
        except ValueError:
            return None

    return None

def _row_contains_tokens(row_vals, tokens):
    """
    tokens: list of lowercase strings to find anywhere in the row
    """
    row_texts = [_norm(v) if _is_str(v) else "" for v in row_vals]
    return all(any(t in cell for cell in row_texts) for t in tokens)

def _find_col_index(row_vals, exact_label):
    """
    Find the column index where the cell equals exact_label (case-insensitive).
    Returns None if not found.
    """
    target = _norm(exact_label)
    for j, v in enumerate(row_vals):
        if _is_str(v) and _norm(v) == target:
            return j
    return None

def _extract_numeric_sequence(row_vals):
    """
    Extract numeric values left->right from a row.
    Returns list[float]
    """
    nums = []
    for v in row_vals:
        val = _parse_money(v)
        if val is not None:
            nums.append(val)
    return nums

def convert_claims_df_to_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Converts your extracted PDF dataframe into:
    Claim Number, Status, Carrier Paid, Carrier Reserve, Carrier Net Incurred
    based on repeated claim blocks.
    """
    out = []
    i = 0
    n = len(df)

    while i < n:
        row = df.iloc[i].tolist()

        # 1) detect the "Claim Number" + "Status" header row
        if _row_contains_tokens(row, ["claim number", "status"]):
            claim_col = _find_col_index(row, "Claim Number")
            status_col = _find_col_index(row, "Status")

            # still try fuzzy
            if claim_col is None:
                claim_col = next((j for j, v in enumerate(row) if _is_str(v) and "claim number" in _norm(v)), None)
            if status_col is None:
                status_col = next((j for j, v in enumerate(row) if _is_str(v) and _norm(v) == "status"), None)

            # next row should have actual values
            if i + 1 < n and claim_col is not None and status_col is not None:
                nxt = df.iloc[i + 1].tolist()
                claim_no = nxt[claim_col] if claim_col < len(nxt) else None
                status = nxt[status_col] if status_col < len(nxt) else None

                # normalize claim number: remove leading zeros but keep as text
                if claim_no is None or (not isinstance(claim_no, str) and pd.isna(claim_no)):
                    claim_no = ""
                else:
                    claim_no = str(claim_no).strip()
                    claim_no = claim_no.lstrip("0") or "0"

                if _is_str(status):
                    status = status.strip()

                # 2) scan forward until we find "Total for Claim:"
                j = i + 2
                total_row_nums = None
                while j < n:
                    rj = df.iloc[j].tolist()
                    if _row_contains_tokens(rj, ["claim number", "status"]):
                        break
                    if any(_is_str(v) and "total for claim" in _norm(v) for v in rj):
                        total_row_nums = _extract_numeric_sequence(rj)
                        break
                    j += 1

                # 3) compute output values from numeric sequence
                loss_paid = total_row_nums[0] if total_row_nums and len(total_row_nums) > 0 else 0.0
                loss_res  = total_row_nums[1] if total_row_nums and len(total_row_nums) > 1 else 0.0
                lae_paid  = total_row_nums[2] if total_row_nums and len(total_row_nums) > 2 else 0.0
                lae_res   = total_row_nums[3] if total_row_nums and len(total_row_nums) > 3 else 0.0
                total_inc = total_row_nums[5] if total_row_nums and len(total_row_nums) > 5 else 0.0

                carrier_paid = loss_paid + lae_paid
                carrier_reserve = loss_res + lae_res
                carrier_net_incurred = total_inc

                out.append({
                    "Claim Number": claim_no,
                    "Status": status,
                    "Carrier Paid": carrier_paid,
                    "Carrier Reserve": carrier_reserve,
                    "Carrier Net Incurred": carrier_net_incurred,
                })

                i = (j + 1) if total_row_nums is not None else (i + 2)
                continue

        i += 1

    return pd.DataFrame(out, columns=[
        "Claim Number", "Status", "Carrier Paid", "Carrier Reserve", "Carrier Net Incurred"
    ])

United_ins_out = convert_claims_df_to_summary(United_ins)
United_ins_out["Status"] = United_ins_out["Status"].astype(str).str.strip().replace(
    {"Closed": "C", "Open": "O", "Reopened": "R"})

# --------------------------------------------------------------------
# 1) RAW TEXT-LAYER WORD + BBOX EXTRACTION - Treo Reports
# --------------------------------------------------------------------
def extract_pdf_text_with_bboxes(pdf_path: str):
    """
    Extract ALL words from PDF with bounding boxes (x0, y0, x1, y1)
    using PyMuPDF text-layer extraction.
    """
    doc = fitz.open(pdf_path)
    words_all = []

    for page_num, page in enumerate(doc, start=1):
        words = page.get_text("words")  
        # Format: [x0, y0, x1, y1, "word", block_no, line_no, span_no]

        for w in words:
            x0, y0, x1, y1, text, block, line, span = w

            words_all.append({
                "page": page_num,
                "text": text,
                "x0": x0,
                "y0": y0,
                "x1": x1,
                "y1": y1
            })

    df = pd.DataFrame(words_all)
    return df


# --------------------------------------------------------------------
# 2) GROUP WORDS INTO ROWS AND COLUMNS
# --------------------------------------------------------------------
def group_ocr_by_row_and_gap(df: pd.DataFrame, y_tolerance: int = 10, column_gap: int = 40):
    """
    Row grouping by y0 proximity,
    Column grouping by horizontal gaps.
    """

    if df is None or df.empty:
        return pd.DataFrame()

    df = df.sort_values(["page", "y0", "x0"]).reset_index(drop=True)
    structured_rows = []

    for page in df["page"].unique():
        page_df = df[df["page"] == page]
        rows = []

        # 1) Group into rows
        for _, r in page_df.iterrows():
            for row in rows:
                if abs(row["y0"] - r["y0"]) <= y_tolerance:
                    row["cells"].append(r)
                    break
            else:
                rows.append({"y0": r["y0"], "cells": [r]})

        # 2) Split into columns by gap
        for row in rows:
            cells = sorted(row["cells"], key=lambda x: x["x0"])

            if not cells:
                structured_rows.append([])
                continue

            columns = []
            current_col = [cells[0]]

            for prev, curr in zip(cells, cells[1:]):
                gap = curr["x0"] - prev["x1"]
                if gap > column_gap:
                    columns.append(current_col)
                    current_col = [curr]
                else:
                    current_col.append(curr)

            columns.append(current_col)

            structured_row = [" ".join(c["text"] for c in col).strip() for col in columns]
            structured_rows.append(structured_row)

    return pd.DataFrame(structured_rows)

# --------------------------------------------------------------------
# 3) MASTER ENTRY POINT — STRUCTURE TEXT-LAYER DATA LIKE OCR OUTPUT
# --------------------------------------------------------------------
def extract_pdf_structured_from_text(pdf_path: str,
                                     y_tolerance: int = 10,
                                     column_gap: int = 40):
    # ✅ Step 1: extract words + bounding boxes
    df_words = extract_pdf_text_with_bboxes(pdf_path)

    # ✅ Step 2: group rows/columns using your OCR algorithm
    df_grouped = group_ocr_by_row_and_gap(
        df_words,
        y_tolerance=y_tolerance,
        column_gap=column_gap
    )
    return df_grouped

# --------------------------------------------------------------------
# MAIN EXECUTION
# --------------------------------------------------------------------
if __name__ == "__main__":
    hartford_path = next(
        (
            p for p in Path(all_input_filepath).glob("*.pdf")
            if any(key in p.name.lower() for key in ["hartford", "treo"])
        ),
        None,
    )

    if hartford_path is None:
        print(f"No PDF containing 'Hartford' found in: {all_input_filepath}")
        time.sleep(15)
        raise FileNotFoundError

    hartford_filename = hartford_path.name
    print("Found Hartford PDF:", hartford_filename)
    hartford_pdf_path = os.path.join(all_input_filepath, hartford_filename)

    # ✅ USE TEXT + BOUNDING BOX PIPELINE (replaces OCR)
    Hartford_df = extract_pdf_structured_from_text(hartford_pdf_path)


Hartford_df.columns = [
    "Claim Number",
    "Claim Status",
    "Accident Date",
    "Claim Description",
    "Total Paid",
    "Indemnity Paid",
    "Indemnity Incurred",
    "Indemnity Outstanding"]

# 4) Optional: tidy column names
Hartford_clean = Hartford_df.copy(deep=True)
Hartford_clean.columns = (
    Hartford_clean.columns.str.strip()
    .str.replace(r"\s+", " ", regex=True)
    .str.replace(r"[/\\]", "-", regex=True))

patterns = ["Policy Period", "Policy Number", "Line of Business", "Claim Number",
            "Claimant Name", "Subtotals", "Grand Totals"]
regex = "|".join(re.escape(p) for p in patterns)

mask = (
    Hartford_clean.iloc[:, :2]
    .astype(str)
    .apply(lambda col: col.str.contains(regex, case=False, na=False))
    .any(axis=1))

Hartford_clean = Hartford_clean.loc[~mask].reset_index(drop=True)
Hartford_clean = Hartford_clean.loc[~Hartford_clean["Total Paid"].isna()].reset_index(drop=True)

def clean_money(col):
    return (
        col.astype(str)
        .str.replace(r"[\$,]", "", regex=True)
        .str.replace(r"\((.*)\)", r"-\1", regex=True)
        .str.replace(" ", "")
        .str.strip())

df = Hartford_clean
if len(df) % 2 != 0:
    df = df.iloc[:-1].copy()

head = df.iloc[0::2].reset_index(drop=True).copy()
tail = df.iloc[1::2].reset_index(drop=True).copy()

head["Total Paid"] = clean_money(head["Total Paid"])
head["Indemnity Outstanding"] = clean_money(head["Indemnity Outstanding"])
tail["Total Paid"] = clean_money(tail["Total Paid"])
tail["Indemnity Outstanding"] = clean_money(tail["Indemnity Outstanding"])

total_paid = pd.to_numeric(head["Total Paid"], errors="coerce")
ind_out = pd.to_numeric(head["Indemnity Outstanding"], errors="coerce")
exp_out = pd.to_numeric(tail["Indemnity Outstanding"], errors="coerce")
total_incurred = pd.to_numeric(tail["Total Paid"], errors="coerce")

Hartford_out2 = pd.DataFrame({
    "Claim Number": head["Claim Number"],
    "Claim Status": head["Claim Status"],
    "Total Paid": total_paid,
    "Total Incurred": total_incurred,
    "Carrier Reserve": ind_out.fillna(0) + exp_out.fillna(0),})

Hartford_out2 = Hartford_out2.reset_index(drop=True)
Hartford_out2["Claim Status"] = Hartford_out2["Claim Status"].astype(str).str.strip().replace(
    {"Closed": "C", "Open": "O", "Reopened": "R"})


# ============================
# CONFIGURATION
# ============================

# --- Source 1 (Inquiry Results) ---
inquiry_path = next((p for p in Path(all_input_filepath).glob("*.xlsx") if "inquiry results" in p.name.lower()), None)
if inquiry_path is None:
    print(f"No Excel file containing 'Inquiry Results' found in: {all_input_filepath}")
    time.sleep(15)
    raise FileNotFoundError(f"No Excel file containing 'Inquiry Results' found in: {all_input_filepath}")
inquiry_filename = inquiry_path.name
print("Found Inquiry Results Excel: ", inquiry_filename)
SOURCE1_XLSX = os.path.join(all_input_filepath, inquiry_filename)
SOURCE1_SHEET_NAME = "Inquiry Results"
SOURCE1_KEY_COL = "Claim Number"

SOURCE1_COLUMN_MAPPING: Dict[str, str] = {
    "Status": "Status: O/C",
    "Paid": "Carrier Paid",
    "Outstanding": "Carrier Reserve",
    "Recovery": "Recovered",
    "Net Incurred": "Carrier Net Incurred",
}

# --- Source 2 (CCMSI) ---
ccmsi_path = next((p for p in Path(all_input_filepath).glob("*.xlsx") if "acme-barricades" in p.name.lower()), None)
if ccmsi_path is None:
    print(f"No Excel file containing 'acme-barricades' found in: {all_input_filepath}")
    time.sleep(15)
    raise FileNotFoundError(f"No Excel file containing 'acme-barricades' found in: {all_input_filepath}")
ccmsi_filename = ccmsi_path.name
print("Found Acme Barriades Monthly Loss runs Excel: ", ccmsi_filename)
SOURCE2_XLSX = os.path.join(all_input_filepath, ccmsi_filename)
SOURCE2_SHEET_NAME = "Sheet3"
SOURCE2_KEY_COL = "Claim Number"

SOURCE2_COLUMN_MAPPING: Dict[str, str] = {
    "Claim Status": "Status: O/C",
    "Total Paid": "Carrier Paid",
    "Outstanding Reserves": "Carrier Reserve",
    "Third Party Recovery": "Recovered",
    "Net Incurred": "Carrier Net Incurred",
}

# --- Source 3 (Loss Runs PDF Input) ---
SOURCE3_KEY_COL = "Claim Number"
SOURCE3_COLUMN_MAPPING: Dict[str, str] = {
    "Status": "Status: O/C",
    "Carrier Paid": "Carrier Paid",
    "Carrier Reserve": "Carrier Reserve",
    "Carrier Net Incurred": "Carrier Net Incurred",
}

# --- Source 4 (Hartford PDF Input) ---
SOURCE4_KEY_COL = "Claim Number"
SOURCE4_COLUMN_MAPPING: Dict[str, str] = {
    "Claim Status": "Status: O/C",
    "Total Paid": "Carrier Paid",
    "Carrier Reserve": "Carrier Reserve",
    "Total Incurred": "Carrier Net Incurred",
}


CLAIMS_XLSX = os.path.join(Claims_filepath, Claims_filename)
CLAIMS_SHEET_NAME = "Auto-WC-GL Claims"
ARCHIVE_SHEET_NAME = "Archive"

# Lookup priority in Claims Log
CLAIMS_KEY_COLS = ["Acme Claim #", "Carrier Claim #", "3rd Party Claim #"]

UPDATE_ALL_MATCHES = True
TREAT_BOTH_BLANKS_AS_EQUAL = True
ROUND_TO_DOLLAR = True
BLANK_EQUALS_ZERO = True

ARCHIVE_HEADERS = [
    "Timestamp", "Source_File", "Target_File", "Sheet",
    "Matched_Column", "Matched_Claim_Number", "Excel_Row",
    "Source_Column", "Target_Column",
    "Old_Value", "New_Value",
]

# ============================
# HELPERS
# ============================
def _clean_header_text(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\u00A0", " ")
    s = s.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    s = s.strip()
    while "  " in s:
        s = s.replace("  ", " ")
    return s

def norm_text(v) -> str:
    if pd.isna(v):
        return ""
    return str(v).strip()

def norm_claim(v) -> str:
    return normalize_for_match(v)

def _strip_common(v: object) -> str:
    if pd.isna(v):
        return ""
    return str(v).replace("\u00A0", " ").strip()

def _remove_numeric_formatting(s: str) -> str:
    return s.replace(",", "").replace("$", "").strip()

def _is_blank_like(v) -> bool:
    if pd.isna(v):
        return True
    s = _strip_common(v)
    if s == "":
        return True
    core = _remove_numeric_formatting(s).replace(" ", "")
    if core in ("", "-", "--"):
        return True
    return False

def _to_decimal_or_none(v) -> Optional[Decimal]:
    if pd.isna(v):
        return None
    s = _strip_common(v)
    if s == "":
        return None
    s_num = _remove_numeric_formatting(s)
    if s_num == "":
        return None
    try:
        return Decimal(s_num)
    except (InvalidOperation, ValueError):
        return None

def _round_dollar_half_up(d: Decimal) -> Decimal:
    return d.quantize(Decimal("1"), rounding=ROUND_HALF_UP)

def parse_numeric_or_text(v):
    if pd.isna(v):
        return v
    s = _strip_common(v)
    if s == "":
        return ""
    s_num = _remove_numeric_formatting(s)
    if s_num == "":
        return ""
    try:
        return float(s_num)
    except Exception:
        return s

def values_equal(a, b) -> bool:
    if TREAT_BOTH_BLANKS_AS_EQUAL and _is_blank_like(a) and _is_blank_like(b):
        return True

    if BLANK_EQUALS_ZERO:
        a_dec0 = _to_decimal_or_none(a)
        b_dec0 = _to_decimal_or_none(b)
        if _is_blank_like(a) and (b_dec0 == Decimal(0) or _is_blank_like(b)):
            return True
        if _is_blank_like(b) and (a_dec0 == Decimal(0) or _is_blank_like(a)):
            return True

    a_dec = _to_decimal_or_none(a)
    b_dec = _to_decimal_or_none(b)
    if a_dec is not None and b_dec is not None:
        if ROUND_TO_DOLLAR:
            return _round_dollar_half_up(a_dec) == _round_dollar_half_up(b_dec)
        else:
            return (a_dec - b_dec).copy_abs() < Decimal("1e-10")

    return _strip_common(a) == _strip_common(b)

def load_source_df(path: str, sheet_name=0) -> pd.DataFrame:
    df = pd.read_excel(path, engine="openpyxl", sheet_name=sheet_name)
    df.columns = [_clean_header_text(c) for c in df.columns]
    return df

def normalize_for_match(v) -> str:
    """
    Canonical form used ONLY for matching
    - strip spaces
    - remove leading zeros
    """
    if v is None or pd.isna(v):
        return ""
    s = str(v).strip()
    s = s.lstrip("0")
    return s or "0"


def normalize_for_excel_text(v) -> str:
    """
    Display form written to Excel
    - preserve leading zeros
    - stored explicitly as TEXT
    """
    if v is None or pd.isna(v):
        return ""
    return str(v).strip()

PANDAS_HEADER_ROW_INDEX = 1
EXCEL_HEADER_ROW_NUMBER = 2 


def load_claims_df(path: str, sheet_name: str) -> pd.DataFrame:
    clean_req = sheet_name.strip()
    xls = pd.ExcelFile(path, engine="openpyxl")
    stripped_map = {name.strip(): name for name in xls.sheet_names}
    resolved_sheet = stripped_map[clean_req]
    df = pd.read_excel(
        path,
        sheet_name=resolved_sheet,
        engine="openpyxl",
        header=PANDAS_HEADER_ROW_INDEX)
    return df

def build_claims_lookup_by_col(
    df_claims: pd.DataFrame,
    claim_key_cols: List[str]
) -> Dict[str, Dict[str, List[int]]]:
    """
    Returns:
      {
        "Acme Claim #": { "123": [row_idx1, row_idx2], ... },
        "Carrier Claim #": { "ABC": [row_idx], ... },
        ...}
    """
    lookups: Dict[str, Dict[str, List[int]]] = {c: defaultdict(list) for c in claim_key_cols}

    for idx, row in df_claims.iterrows():
        for col in claim_key_cols:
            if col not in df_claims.columns:
                continue
            val = norm_claim(row.get(col, ""))
            if val:
                lookups[col][val].append(idx)

    # de-dupe indices per (col, claim)
    for col in lookups:
        for claim_val, idxs in list(lookups[col].items()):
            lookups[col][claim_val] = list(dict.fromkeys(idxs))
    return lookups

def get_excel_row_index_from_df_index(df_index: int) -> int:
    return EXCEL_HEADER_ROW_NUMBER + 1 + df_index

def get_column_index_map(ws) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    max_col = ws.max_column or 0
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=EXCEL_HEADER_ROW_NUMBER, column=col_idx)
        name = _clean_header_text(cell.value)
        if name:
            col_map[name] = col_idx
    return col_map

def ensure_archive_sheet(wb):
    if ARCHIVE_SHEET_NAME in wb.sheetnames:
        ws_arc = wb[ARCHIVE_SHEET_NAME]
    else:
        ws_arc = wb.create_sheet(ARCHIVE_SHEET_NAME)

    first_row_values = [
        _clean_header_text(ws_arc.cell(row=1, column=j).value)
        for j in range(1, len(ARCHIVE_HEADERS) + 1)
    ]
    expected = [_clean_header_text(h) for h in ARCHIVE_HEADERS]

    if first_row_values != expected:
        for j, header in enumerate(ARCHIVE_HEADERS, start=1):
            ws_arc.cell(row=1, column=j, value=header)

    return ws_arc

# ==========================================================
# ✅ CLEAR PRIOR ROW HIGHLIGHTS
# ==========================================================

def clear_row_highlights(ws, header_row: int):
    """
    Clears background fill from all data rows
    """
    max_row = ws.max_row or header_row
    max_col = ws.max_column or 0

    for r in range(header_row + 1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)

# ============================
# SYNC LOGIC FOR ONE SOURCE (DF-BASED)
# ============================
def sync_one_source_df(
    src_df: pd.DataFrame,
    src_key_col: str,
    col_mapping: Dict[str, str],
    clm_df: pd.DataFrame,
    claims_lookup: Dict[str, Dict[str, List[int]]],  # lookup-by-column
    claim_key_cols: List[str],
    ws_claims,
    ws_arc,
    target_col_to_idx: Dict[str, int],
    claims_sheet: str,
    source_label: str,
    claims_xlsx: str,
) -> Tuple[int, int]:

    src_df = src_df.copy()
    src_df.columns = [_clean_header_text(c) for c in src_df.columns]

    src_key = _clean_header_text(src_key_col)
    col_mapping = {_clean_header_text(k): _clean_header_text(v) for k, v in col_mapping.items()}

    missing_in_source = [c for c in [src_key] + list(col_mapping.keys()) if c not in src_df.columns]
    if missing_in_source:
        print("DEBUG: Columns read from source:", [repr(c) for c in src_df.columns])
        raise KeyError(f"Source '{source_label}' missing columns: {missing_in_source}")

    total_changes = 0
    archive_rows_buffer = []

    source_file_name = source_label
    target_file_name = os.path.basename(claims_xlsx)
    timestamp = datetime.now().strftime("%Y-%m-%d")

    # ✅ NEW: prevent multiple updates to the same Claims Log row from the same source run
    processed_clm_idxs = set()
    changed_excel_rows = set()

    # ✅ NEW: process in priority order (Acme -> Carrier -> 3rd Party)
    for key_col in claim_key_cols:
        if key_col not in claims_lookup:
            continue

        for _, s_row in src_df.iterrows():
            key_val_src = norm_claim(s_row.get(src_key, ""))
            if not key_val_src:
                continue

            # Only consider matches for THIS key_col in this pass
            matched_rows = claims_lookup[key_col].get(key_val_src, [])
            if not matched_rows:
                continue

            if not UPDATE_ALL_MATCHES:
                matched_rows = matched_rows[:1]

            for clm_idx in matched_rows:
                # ✅ skip if this Claims Log row already updated by a higher-priority match
                if clm_idx in processed_clm_idxs:
                    continue

                excel_row = get_excel_row_index_from_df_index(clm_idx)

                for src_col, tgt_col in col_mapping.items():
                    src_val_raw = s_row.get(src_col, None)
                    tgt_val_raw = clm_df.at[clm_idx, tgt_col]

                    src_val = parse_numeric_or_text(src_val_raw)
                    tgt_val = parse_numeric_or_text(tgt_val_raw)

                    if values_equal(src_val, tgt_val):
                        continue

                    clm_df.at[clm_idx, tgt_col] = src_val
                    ws_claims.cell(row=excel_row, column=target_col_to_idx[tgt_col], value=src_val)
                    changed_excel_rows.add(excel_row)

                    archive_rows_buffer.append([
                        timestamp,
                        source_file_name,
                        target_file_name,
                        claims_sheet,
                        key_col,          # matched column (this pass)
                        key_val_src,
                        excel_row,
                        src_col,
                        tgt_col,
                        tgt_val_raw,
                        src_val_raw
                    ])
                    total_changes += 1

                # ✅ mark as processed AFTER applying updates for this source row
                processed_clm_idxs.add(clm_idx)

    for row in archive_rows_buffer:
        ws_arc.append(row)

    return total_changes, len(archive_rows_buffer), changed_excel_rows

# ============================
# MAIN: SOURCES 1 + 2 (Excel) + 3 (PDF) + 4 (PDF)
# ============================
def sync_sources_to_claims(
    source3_df: pd.DataFrame,
    source3_label: str = "Source3 - United Insurance PDF",
    source4_df: Optional[pd.DataFrame] = None,
    source4_label: str = "Source4 - Hartford PDF"
):
    """
    source3_df columns:
      Claim Number, Status, Carrier Paid, Carrier Reserve, Carrier Net Incurred

    source4_df (Hartford_out2) columns:
      Claim Number, Claim Status, Total Paid, Carrier Reserve, Total Incurred
    """
    print("Loading Claims dataframe...")
    clm_df = load_claims_df(CLAIMS_XLSX, CLAIMS_SHEET_NAME)
    claim_key_cols = [_clean_header_text(c) for c in CLAIMS_KEY_COLS]
    if not any(c in clm_df.columns for c in claim_key_cols):
        raise KeyError(f"None of the CLAIMS_KEY_COLS were found in Claims sheet columns: {claim_key_cols}")

    # ✅ FIX: Use lookup-by-column so we can prioritize matching and avoid duplicates
    claims_lookup = build_claims_lookup_by_col(clm_df, claim_key_cols)

    wb = load_workbook(CLAIMS_XLSX)
    stripped_map = {name.strip(): name for name in wb.sheetnames}
    clean_name = CLAIMS_SHEET_NAME.strip()
    if clean_name not in stripped_map:
        raise KeyError(f"Sheet '{CLAIMS_SHEET_NAME}' not found in Claims workbook.")
    ws_claims = wb[stripped_map[clean_name]]


    col_map = get_column_index_map(ws_claims)
    ws_arc = ensure_archive_sheet(wb)

    all_target_cols = (
        set(SOURCE1_COLUMN_MAPPING.values())
        | set(SOURCE2_COLUMN_MAPPING.values())
        | set(SOURCE3_COLUMN_MAPPING.values())
        | set(SOURCE4_COLUMN_MAPPING.values())
    )

    target_col_to_idx: Dict[str, int] = {}
    for tgt_col in all_target_cols:
        tgt_col_clean = _clean_header_text(tgt_col)
        if tgt_col_clean not in col_map:
            raise KeyError(
                f"Target column '{tgt_col_clean}' not found in '{CLAIMS_SHEET_NAME}'. "
                f"Available: {list(col_map.keys())}"
            )
        target_col_to_idx[tgt_col_clean] = col_map[tgt_col_clean]

    # Load Excel sources
    print("Loading Inquiry Results Excel from sheet 'Inquiry Results'...")
    src1_df = load_source_df(SOURCE1_XLSX, sheet_name=SOURCE1_SHEET_NAME)
    src1_df["Status"] = src1_df["Status"].astype(str).str.strip().replace({"Closed": "C", "Open": "O", "Reopened": "R"})

    print("Loading Acme Barricades Monthly Loss Runs Excel from sheet 'Sheet3'...")
    src2_df = load_source_df(SOURCE2_XLSX, sheet_name=SOURCE2_SHEET_NAME)
    src2_df["Claim Status"] = src2_df["Claim Status"].astype(str).str.strip().replace({"Closed": "C", "Open": "O", "Reopened": "R"})

    # --- Sync Source 1 ---
    print(f"\n--- Syncing Source 1: {os.path.basename(SOURCE1_XLSX)} ---")
    changes_1, arc_1, rows_1 = sync_one_source_df(
        src_df=src1_df,
        src_key_col=SOURCE1_KEY_COL,
        col_mapping=SOURCE1_COLUMN_MAPPING,
        clm_df=clm_df,
        claims_lookup=claims_lookup,
        claim_key_cols=claim_key_cols,
        ws_claims=ws_claims,
        ws_arc=ws_arc,
        target_col_to_idx=target_col_to_idx,
        claims_sheet=CLAIMS_SHEET_NAME,
        source_label=os.path.basename(SOURCE1_XLSX),
        claims_xlsx=CLAIMS_XLSX
    )
    print(f"Source 1 updates: {changes_1} | Archive rows: {arc_1}")

    # --- Sync Source 2 ---
    print(f"\n--- Syncing Source 2: {os.path.basename(SOURCE2_XLSX)} ---")
    changes_2, arc_2, rows_2 = sync_one_source_df(
        src_df=src2_df,
        src_key_col=SOURCE2_KEY_COL,
        col_mapping=SOURCE2_COLUMN_MAPPING,
        clm_df=clm_df,
        claims_lookup=claims_lookup,
        claim_key_cols=claim_key_cols,
        ws_claims=ws_claims,
        ws_arc=ws_arc,
        target_col_to_idx=target_col_to_idx,
        claims_sheet=CLAIMS_SHEET_NAME,
        source_label=os.path.basename(SOURCE2_XLSX),
        claims_xlsx=CLAIMS_XLSX
    )
    print(f"Source 2 updates: {changes_2} | Archive rows: {arc_2}")

    # --- Sync Source 3 ---
    print(f"\n--- Syncing Source 3 (DF): {source3_label} ---")
    changes_3, arc_3, rows_3 = sync_one_source_df(
        src_df=source3_df,
        src_key_col=SOURCE3_KEY_COL,
        col_mapping=SOURCE3_COLUMN_MAPPING,
        clm_df=clm_df,
        claims_lookup=claims_lookup,
        claim_key_cols=claim_key_cols,
        ws_claims=ws_claims,
        ws_arc=ws_arc,
        target_col_to_idx=target_col_to_idx,
        claims_sheet=CLAIMS_SHEET_NAME,
        source_label=source3_label,
        claims_xlsx=CLAIMS_XLSX
    )
    print(f"Source 3 updates: {changes_3} | Archive rows: {arc_3}")

    # --- Sync Source 4 ---
    changes_4, arc_4, rows_4 = 0, 0, 0
    if source4_df is not None and not source4_df.empty:
        print(f"\n--- Syncing Source 4 (DF): {source4_label} ---")
        changes_4, arc_4, rows_4 = sync_one_source_df(
            src_df=source4_df,
            src_key_col=SOURCE4_KEY_COL,
            col_mapping=SOURCE4_COLUMN_MAPPING,
            clm_df=clm_df,
            claims_lookup=claims_lookup,
            claim_key_cols=claim_key_cols,
            ws_claims=ws_claims,
            ws_arc=ws_arc,
            target_col_to_idx=target_col_to_idx,
            claims_sheet=CLAIMS_SHEET_NAME,
            source_label=source4_label,
            claims_xlsx=CLAIMS_XLSX
        )
        print(f"Source 4 updates: {changes_4} | Archive rows: {arc_4}")
    else:
        print("\n--- Source 4 skipped (empty or None) ---")

    # ==========================================================
    # ✅ COLLECT ALL CHANGED ROWS (FROM ALL SOURCES)
    # ==========================================================

    all_changed_rows = set()

    for s in [rows_1, rows_2, rows_3, rows_4]:
        if s:
            all_changed_rows.update(s)
    # ==========================================================
    # ✅ BUILD LIST OF ALL CLAIMS FOUND IN ALL 4 INPUT SOURCES
    # ==========================================================
    print("\nCollecting unmatched claims...")

    # Normalize helper
    def _norm_claim_id(x):
        return normalize_for_match(x)

    # Collect claim numbers from each source
    src1_claims = set(_norm_claim_id(c) for c in src1_df[SOURCE1_KEY_COL].astype(str))
    src2_claims = set(_norm_claim_id(c) for c in src2_df[SOURCE2_KEY_COL].astype(str))
    src3_claims = set(_norm_claim_id(c) for c in source3_df[SOURCE3_KEY_COL].astype(str))
    src4_claims = set(_norm_claim_id(c) for c in source4_df[SOURCE4_KEY_COL].astype(str)) if source4_df is not None else set()

    all_input_claims = src1_claims | src2_claims | src3_claims | src4_claims

    # Now get ALL claim numbers from Claims Log (Acme/Carrier/3rd Party)
    all_claims_log = set()
    for col in claim_key_cols:
        if col in clm_df.columns:
            all_claims_log |= set(_norm_claim_id(v) for v in clm_df[col].astype(str))

    # ✅ Find claim numbers that appear in inputs but NOT found anywhere in Claims Log
    unmatched_claims = sorted(all_input_claims - all_claims_log)

    print(f"\nUnmatched claims found: {len(unmatched_claims)}")

    # ==========================================================
    # ✅ BUILD OUTPUT DATAFRAME FOR THESE CLAIMS & FINANCES
    # ==========================================================
    not_matched_records = []

    # Helper to fetch financials from ANY source df
    def find_in_sources(claim_no):
        claim_no_norm = _norm_claim_id(claim_no)

        # search in source1
        row = src1_df[src1_df[SOURCE1_KEY_COL].astype(str).apply(_norm_claim_id) == claim_no_norm]
        if not row.empty:
            r = row.iloc[0]
            orig_claim = r.get(SOURCE1_KEY_COL)
            return {
                "Claim Number": normalize_for_excel_text(orig_claim),
                "Carrier Paid": r.get("Carrier Paid", None),
                "Carrier Reserve": r.get("Carrier Reserve", None),
                "Carrier Net Incurred": r.get("Carrier Net Incurred", None),
                "Found In": "Inquiry Results"
            }

        # search in source2
        row = src2_df[src2_df[SOURCE2_KEY_COL].astype(str).apply(_norm_claim_id) == claim_no_norm]
        if not row.empty:
            r = row.iloc[0]
            orig_claim = r.get(SOURCE2_KEY_COL)
            return {
                "Claim Number": normalize_for_excel_text(orig_claim),
                "Carrier Paid": r.get("Carrier Paid", None),
                "Carrier Reserve": r.get("Carrier Reserve", None),
                "Carrier Net Incurred": r.get("Carrier Net Incurred", None),
                "Found In": "CCMSI Monthly Loss Runs"
            }

        # search in source3
        row = source3_df[source3_df[SOURCE3_KEY_COL].astype(str).apply(_norm_claim_id) == claim_no_norm]
        if not row.empty:
            r = row.iloc[0]
            orig_claim = r.get(SOURCE3_KEY_COL)
            return {
                "Claim Number": normalize_for_excel_text(orig_claim),
                "Carrier Paid": r.get("Carrier Paid", None),
                "Carrier Reserve": r.get("Carrier Reserve", None),
                "Carrier Net Incurred": r.get("Carrier Net Incurred", None),
                "Found In": "United Insurance Loss Runs PDF"
            }

        # search in source4
        if source4_df is not None:
            row = source4_df[source4_df[SOURCE4_KEY_COL].astype(str).apply(_norm_claim_id) == claim_no_norm]
            if not row.empty:
                r = row.iloc[0]
                orig_claim = r.get(SOURCE4_KEY_COL)
                return {
                    "Claim Number": normalize_for_excel_text(orig_claim),
                    "Carrier Paid": r.get("Total Paid", None),
                    "Carrier Reserve": r.get("Carrier Reserve", None),
                    "Carrier Net Incurred": r.get("Total Incurred", None),
                    "Found In": "Hartford Loss Runs PDF"
                }

        return {"Claim Number": normalize_for_excel_text(claim_no), "Carrier Paid": None, "Carrier Reserve": None,
                "Carrier Net Incurred": None, "Found In": "Unknown"}

    # Build list of dict rows
    for claim_no in unmatched_claims:
        not_matched_records.append(find_in_sources(claim_no))

    df_unmatched = pd.DataFrame(not_matched_records)

    # ==========================================================
    # ✅ ADD "CLAIMS NOT MATCHED" TAB INTO CLAIMS LOG EXCEL
    # ==========================================================

    df_unmatched_final = df_unmatched.copy()

    # Add run date
    df_unmatched_final.insert(
        0,
        "Updated On",
        datetime.now().strftime("%Y-%m-%d")
    )

    # Rename key column for business users
    df_unmatched_final = df_unmatched_final.rename(
        columns={"Claim Number": "Acme Claim #"}
    )

    # Reorder columns
    df_unmatched_final = df_unmatched_final[
        [
            "Updated On",
            "Acme Claim #",
            "Carrier Paid",
            "Carrier Reserve",
            "Carrier Net Incurred",
            "Found In",
        ]
    ]

    UNMATCHED_SHEET_NAME = "Claims not matched"

    # Remove existing tab if present (refresh each run)
    if UNMATCHED_SHEET_NAME in wb.sheetnames:
        del wb[UNMATCHED_SHEET_NAME]

    ws_unmatched = wb.create_sheet(UNMATCHED_SHEET_NAME)

    # Write header
    for col_idx, col_name in enumerate(df_unmatched_final.columns, start=1):
        ws_unmatched.cell(row=1, column=col_idx, value=col_name)

    # Write rows
    for row_idx, row in enumerate(
        df_unmatched_final.itertuples(index=False), start=2
    ):
        for col_idx, value in enumerate(row, start=1):
            cell = ws_unmatched.cell(row=row_idx, column=col_idx)
            if df_unmatched_final.columns[col_idx - 1] == "Acme Claim #":
                cell.value = str(value)
                cell.number_format = "@"
            else:
                cell

    print("✅ 'Claims not matched' tab added to Claims Log")
    # ==========================================================
    # ✅ CLEAR OLD HIGHLIGHTS & APPLY NEW ONES
    # ==========================================================

    print("\n---Clearing prior highlights...")
    clear_row_highlights(ws_claims, EXCEL_HEADER_ROW_NUMBER)

    print(f"Highlighting {len(all_changed_rows)} updated rows...")

    for excel_row in all_changed_rows:
        for col in range(1, ws_claims.max_column + 1):
            ws_claims.cell(
                row=excel_row,
                column=col
            ).fill = CHANGE_HIGHLIGHT_FILL
    wb.save(CLAIMS_XLSX)
    print("\n✅ Completed sources 1, 2, 3, and 4.")
    print(f"Total cell updates applied: {changes_1 + changes_2 + changes_3 + changes_4}")
    print(f"Total archive rows appended: {arc_1 + arc_2 + arc_3 + arc_4}")
    print(f"Saved: {CLAIMS_XLSX}")

if __name__ == "__main__":
    source3_df = United_ins_out
    source3_label = "LossRunsReport.pdf"

    source4_df = Hartford_out2
    source4_label = "Acme Hartford CGL Loss Runs - monthly.pdf"

    sync_sources_to_claims(
        source3_df=source3_df,
        source3_label=source3_label,
        source4_df=source4_df,
        source4_label=source4_label
    )